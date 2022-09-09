import datetime
import os

import openpyxl
import pathlib

from openpyxl.styles import Font

workbook = openpyxl.load_workbook('etual.xlsx')
masters = workbook['Производство']


def plus_cabinet(cabinet, name, mass):

    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    names = return_row("Производство", "Наименование")
    cabinets = return_cabinet()
    adress_cab = None
    simvol = None
    simvols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    for i in range(0, len(simvols)):
        if data[str(simvols[i])+"1"].value == cabinet:
            simvol = simvols[i]

    for i in range(0, len(names)):
        if names[i] == name:
            if data[str(simvol) + str(i + 2)].value == None:
                data[str(simvol) + str(i + 2)].value = + int(mass)
            else:
                data[str(simvol) + str(i+2)].value = int(data[str(simvol) + str(i+2)].value) + int(mass)
    workbook_temp.save('etual.xlsx')




def return_xyz(name):
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    for i in range(1, 1000):
        if data['D' + str(i)].value == name:
            return data['E'+str(i)].value

def brend_for_cosmetic(cosmetic, question):
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    for i in range(1, 1000):
        if data['D'+str(i)].value == cosmetic:
            if question == "brend":
                return data['C'+str(i)].value
            if question == "my_price":
                return data['G'+str(i)].value/data['E'+str(i)].value


def new_cosmetic(record):
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    new_row_index = None
    simvols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    for i in range(1, 10000):
        if str(data['D' + str(i)].value) == "None":
            new_row_index = i
            break
    data['A'+str(new_row_index)] = int(new_row_index)-1

    for r in range(0, len(record)):
        data[str(simvols[r+1])+str(new_row_index)] = record[r]
    workbook_temp.save('etual.xlsx')


def add_cosmetic(record):
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    for i in range(1,1000):
            if data['D'+str(i)].value == str(record[1]):
                if data['I'+str(i)].value != None:
                    data['I'+str(i)]=int(data['I'+str(i)].value)+int(record[2])
                else:
                    data['I' + str(i)] = int(record[2])
                data['G' + str(i)] = int(record[3])
    workbook_temp.save('etual.xlsx')


def return_name(adress):
    try:
        list = str(adress[1:adress.index("!")])
    except:
        return "error"
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    masters_temp = workbook_temp[list]
    result = str(masters_temp[adress[adress.index("!") + 1:]].value)
    return result


def searh(means, cabinet):  #поиск названия по ссылке в Ексель
    data = None
    # Iterate the loop to read the cell values
    for i in range(0, masters.max_row):
        for col in masters.iter_cols(0, masters.max_column):
            if str(col[i].value)[0] == "=":
                print(return_name(str(col[i].value)), end="\t\t")
            else:
                print(col[i].value, end="\t\t")
    return data


def return_cabinet():
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    result = []
    simvols = ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'R', 'S', 'T', 'U', 'V']
    for i in range(0, len(simvols)):
        if data[simvols[i] + "1"].value is not None:
            result.append(data[simvols[i]+"1"].value)
    return result


def return_row(name_list, name_column): #возвращает данные колонки
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp[name_list]
    result = []
    simvols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    coord = None
    for i in range(1, 1000):
        if data[str(simvols[i]) + '1'].value == name_column:
            coord = simvols[i]
            break
    for i in range(2, 1000):
        if data[str(coord) + str(i)].value != None:
            result.append(str(data[str(coord) + str(i)].value))
        else:
            result.append(0)
    return result


def return_cabitet_cosmetic(cabinet):
    cosmetic_name = return_row("Производство", "Наименование")
    cosmetic_weight = return_row("Производство", cabinet)
    return dict(zip(cosmetic_name, cosmetic_weight))


def save_in_file(name, meaning, meaning_now, cabinet, master):
    workbook_temp = openpyxl.load_workbook('etual.xlsx')
    data = workbook_temp["Производство"]
    simvols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']
    coord = None
    for i in range(0, 1000):
        if data[str(simvols[i]) + '1'].value == cabinet:
            coord = simvols[i]
            break
    for j in range(0, len(name)):
        for i in range(1, 1000):
            if name[j] == data['D'+str(i)].value:
                data[coord + str(i)] = int(meaning_now[j])
    workbook_temp.save('etual.xlsx')


    workbook_temp = openpyxl.load_workbook('shablon.xlsx')
    data = workbook_temp["Шаблон"]
    data['B4'] = master
    data['B2'] = cabinet
    row = 7
    itog = 0
    otchet = []
    itog_brand = 0
    for i in range(0, len(meaning)):
        if meaning[i] != meaning_now[i]:
            otchet.append([(brend_for_cosmetic(name[i], "brend")), name[i], int(meaning[i]) - int(meaning_now[i]), float('{:.2f}'.format(brend_for_cosmetic(name[i], "my_price")*(float(meaning[i]) - float(meaning_now[i]))))])
    otchet.sort(key=lambda x:x[0], reverse=False)
    for i in range(len(otchet)):
        print(otchet[i][3])
    x = len(otchet)
    for i in range(0, x):
        if i != 0 and otchet[i][0] != otchet[i-1][0]:
            data["E" + str(row)] = itog_brand
            data["E" + str(row)].font = Font(bold=True, size=14)
            data["D" + str(row)] = "Итого"
            data["D" + str(row)].font = Font(bold=True, size=14)
            itog_brand = 0
            row = row + 2
        data["B" + str(row)] = otchet[i][0]
        data["C" + str(row)] = otchet[i][1]
        data["D"+str(row)] = otchet[i][2]
        data["E" + str(row)] = otchet[i][3]
        itog = float(data["E" + str(row)].value + itog)
        itog_brand = itog_brand + otchet[i][3]
        row = row+1

    data["E" + str(row)] = itog_brand
    data["E" + str(row)].font = Font(bold=True, size=14)
    data["D" + str(row)] = "Итого"
    data["D" + str(row)].font = Font(bold=True, size=14)
    itog_brand = 0


    data["E" + str(row+1)] = itog
    data["E" + str(row + 1)].font = Font(bold=True, size=14)
    data["D" + str(row+1)] = "Итого расход"
    data["D" + str(row+1)].font = Font(bold=True, size=14)
    dir_path = pathlib.Path.cwd()
    data["B3"] = str(datetime.date.today() - datetime.timedelta(days=1))
    try:
        os.makedirs('history' +"\\"+ str(datetime.date.today()) + '')
    except: pass
    workbook_temp.save(pathlib.Path(dir_path, 'history' +"\\"+ str(datetime.date.today()) + '', '' + cabinet + '.xlsx'))


if __name__ == "__main__":
    pass
